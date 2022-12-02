from calendar import month
from datetime import date, datetime
from urllib.request import Request
from django.http import HttpResponseNotAllowed
from django.shortcuts import redirect, render, get_object_or_404, redirect
from django.http import HttpResponse
from .forms import Form_Quarta, Form_Sabado
from dateutil import relativedelta
from django.contrib import messages
from django.core.paginator import Paginator
from django.contrib.auth.decorators import login_required

from .models import DesigQuarta, DesigSabado
from openpyxl import load_workbook
import pandas as pd 



def home(request):
    return render(request, 'cp/home.html') 

def mapa(request):
    return render(request, 'cp/mapa.html')

def escritorio(request):
    return render(request, 'cp/escritorio.html')

def fotos(request):
    return render(request, 'cp/fotos.html')

def local(request):
    return render(request, 'cp/local.html')

def limpeza(request):
    return render(request, 'cp/limpeza.html')

def reunioes(request):
    return render(request, 'cp/reunioes.html')

def congresso(request):
    return render(request, 'cp/congresso.html')

def grupos(request):
    return render(request, 'cp/grupos.html')

def relatorio(request):
    return render(request, 'cp/relatório.html')

def visita(request):
    return render(request, 'cp/visita.html')   

@login_required
def restrita(request):
    return render(request, 'cp/restrita.html')                                                    

@login_required
def desig_quarta(request):
    if request.method == 'POST':
        quarta = Form_Quarta(request.POST)
         
        if quarta.is_valid() :
            quarta.save()
            messages.info(request, 'Adicionado com sucesso')
            quarta = Form_Quarta()
            return redirect('/historico_quarta')
    
    else:
        quarta = Form_Quarta()
       
    return render(request, 'cp/desig_quarta.html', {'quarta': quarta})             

@login_required
def desig_sabado(request):
    if request.method == 'POST':
        sabado = Form_Sabado(request.POST)
        
        if sabado.is_valid():
            sabado.save()
            messages.info(request, 'Adicionado com sucesso')
            sabado = Form_Sabado()
            return redirect('/historico_sabado')
   
    else:
        sabado = Form_Sabado()

    return render(request, 'cp/desig_sabado.html', {'sabado': sabado}) 

@login_required
def historico_quarta(request):
    mes = request.GET.get('data')
    quarta = DesigQuarta.objects.order_by('data').all()
    
   

    if mes:
        quarta = DesigQuarta.objects.filter(data__icontains=mes)
        imprimir = DesigQuarta.objects.filter(data__icontains=mes).values_list()
        df = pd.DataFrame(imprimir)
        df1 = df.transpose()
        designacao = load_workbook("static/Designações_Padrão.xlsx")
        writer = pd.ExcelWriter("static/Designações_Mensal.xlsx")
        writer.book = designacao
        writer.sheets = dict((ws.title, ws) for ws in designacao.worksheets)
        df1.to_excel(writer, 'HOME', startrow=2, startcol=5, header=False, index=False)
        writer.save()
       
    else:
        return render(request, 'cp/historico_quarta.html', {'quarta': quarta, 'indicadores': indicadores})

    return render(request, 'cp/historico_quarta.html',  {'quarta': quarta, 'indicadores': indicadores})  

@login_required
def historico_sabado(request):
    mes = request.GET.get('data')
    sabado = DesigSabado.objects.order_by('data').all()
   

    if mes:
        sabado = DesigSabado.objects.filter(data__icontains=mes)
        imprimir = DesigSabado.objects.filter(data__icontains=mes).values_list()
        df = pd.DataFrame(imprimir)
        df1 = df.transpose()
        designacao = load_workbook("static/Designações_Mensal.xlsx")
        writer = pd.ExcelWriter("static/Designações_Mensal.xlsx")
        writer.book = designacao
        writer.sheets = dict((ws.title, ws) for ws in designacao.worksheets)
        df1.to_excel(writer, 'HOME', startrow=27, startcol=5, header=False, index=False)
        writer.save()
     
       
    else:
        return render(request, 'cp/historico_sabado.html', {'sabado': sabado})

    return render(request, 'cp/historico_sabado.html', {'sabado': sabado})  
    
@login_required    
def editar_quarta(request, id):
    quarta = get_object_or_404(DesigQuarta, pk=id)
    form = Form_Quarta(instance=quarta)
 
    if request.method == 'POST':
        form = Form_Quarta(request.POST, instance=quarta)
         
        if form.is_valid():
            quarta.save()
            messages.info(request, 'Editado com sucesso')
            return redirect('/historico_quarta')
            
   
        else:
            return render(request, 'cp/edit_quarta.html', {'form':form ,'quarta': quarta})  


    return render(request, 'cp/edit_quarta.html', {'form':form ,'quarta': quarta})  

@login_required
def deletar_quarta(request, id):
    del_quarta = get_object_or_404(DesigQuarta, pk=id)
   
    if request.method == 'POST':
        del_quarta.delete()

        messages.info(request, 'Apagado com sucesso')
        return redirect('/historico_quarta')

    return render(request, 'cp/deletar_quarta.html')

@login_required
def editar_sabado(request, id):
    sabado = get_object_or_404(DesigSabado, pk=id)
    form = Form_Sabado(instance=sabado)
 
    if request.method == 'POST':
        form = Form_Sabado(request.POST, instance=sabado)
         
        if form.is_valid():
            sabado.save()
            messages.info(request, 'Editado com sucesso')
            return redirect('/historico_sabado')
   
        else:
            return render(request, 'cp/edit_sabado.html', {'form':form ,'sabado': sabado})  


    return render(request, 'cp/edit_sabado.html', {'form':form ,'sabado': sabado})  

@login_required
def deletar_sabado(request, id):
    del_sabado = get_object_or_404(DesigSabado, pk=id)
   
    if request.method == 'POST':
        del_sabado.delete()

        messages.info(request, 'Apagado com sucesso')
        return redirect('/historico_sabado')

    return render(request, 'cp/deletar_sabado.html')

def indicadores(request):
    if request.GET.get('data'):
        mes = request.GET.get('data')
        quarta = DesigQuarta.objects.filter(data__icontains=mes).order_by('data')
        sabado = DesigSabado.objects.filter(data__icontains=mes).order_by('data')

    else:
        dt = date.today()
        dt1 = dt + relativedelta.relativedelta(months=1)
        mes = dt1.strftime('%Y-%m')
        quarta = DesigQuarta.objects.filter(data__icontains=mes).order_by('data')
        sabado = DesigSabado.objects.filter(data__icontains=mes).order_by('data')

   
    return render(request, 'cp/indicadores.html', {'quarta': quarta, 'sabado': sabado})

def indicador_quarta(request, id):
    quarta = get_object_or_404(DesigQuarta, pk=id)
    form = Form_Quarta(instance=quarta)
 
    if request.method == 'POST':
        form = Form_Quarta(request.POST, instance=quarta)
         
        if form.is_valid():
            quarta.save()
           
            messages.info(request, 'Editado com sucesso')
            return redirect('/indicadores')
            
   
        else:
            return render(request, 'cp/indicador_quarta.html', {'form':form ,'quarta': quarta})  


    return render(request, 'cp/indicador_quarta.html', {'form':form ,'quarta': quarta})  

def indicador_sabado(request, id):
    sabado = get_object_or_404(DesigSabado, pk=id)
    form = Form_Sabado(instance=sabado)
 
    if request.method == 'POST':
        form = Form_Sabado(request.POST, instance=sabado)
         
        if form.is_valid():
            sabado.save()
            messages.info(request, 'Editado com sucesso')
            return redirect('/indicadores')
            
   
        else:
            return render(request, 'cp/indicador_sabado.html', {'form':form ,'sabado': sabado})  


    return render(request, 'cp/indicador_sabado.html', {'form':form ,'sabado': sabado})  



def desig_coordenador(request):
    if request.GET.get('data'):
        mes = request.GET.get('data')
        quarta = DesigQuarta.objects.filter(data__icontains=mes).order_by('data')
        
    else:
        dt = date.today()
        dt1 = dt + relativedelta.relativedelta(months=1)
        mes = dt1.strftime('%Y-%m')
        quarta = DesigQuarta.objects.filter(data__icontains=mes).order_by('data')
        

   
    return render(request, 'cp/desig_coordenador.html', {'quarta': quarta})

def desig_coordenador_quarta(request, id):
    quarta = get_object_or_404(DesigQuarta, pk=id)
    form = Form_Quarta(instance=quarta)
 
    if request.method == 'POST':
        form = Form_Quarta(request.POST, instance=quarta)
         
        if form.is_valid():
            quarta.save()
            messages.info(request, 'Editado com sucesso')
            return redirect('/desig_coordenador')
            
   
        else:
            return render(request, 'cp/desig_coordenador_quarta.html', {'form':form ,'quarta': quarta})  


    return render(request, 'cp/desig_coordenador_quarta.html', {'form':form ,'quarta': quarta}) 

def desig_conferencia(request):
    if request.GET.get('data'):
        mes = request.GET.get('data')
     
        sabado = DesigSabado.objects.filter(data__icontains=mes).order_by('data')

    else:
        dt = date.today()
        dt1 = dt + relativedelta.relativedelta(months=1)
        mes = dt1.strftime('%Y-%m')

        sabado = DesigSabado.objects.filter(data__icontains=mes).order_by('data')

   
    return render(request, 'cp/desig_conferencia.html', {'sabado': sabado})



def desig_conferencia2(request, id):
    sabado = get_object_or_404(DesigSabado, pk=id)
    form = Form_Sabado(instance=sabado)
 
    if request.method == 'POST':
        form = Form_Sabado(request.POST, instance=sabado)
         
        if form.is_valid():
            sabado.save()
            messages.info(request, 'Editado com sucesso')
            return redirect('/desig_conferencia')
            
   
        else:
            return render(request, 'cp/desig_conferencia2.html', {'form':form ,'sabado': sabado})  


    return render(request, 'cp/desig_conferencia2.html', {'form':form ,'sabado': sabado})  




